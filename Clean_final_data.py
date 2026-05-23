# -*- coding: utf-8 -*-

from openpyxl import load_workbook

excel_path = 'TL/final_data.xlsx'
wb = load_workbook(excel_path)
keep_sheet_name = 'Cycle 0'
sheets = wb.sheetnames
for sheet in sheets:
    if sheet != keep_sheet_name:
        wb.remove(wb[sheet])
wb.save(excel_path)