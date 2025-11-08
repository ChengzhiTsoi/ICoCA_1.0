# -*- coding: utf-8 -*-

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# Read txt file
with open('N.txt') as f:
    lines = f.readlines()
num_lines = len(lines)
i = 0

# Read Excel file
excel_path = 'TL_data_target_task_test.xlsx'
df = pd.read_excel(excel_path, header = None, engine = 'openpyxl')
row_count = len(df)
col_count = df.shape[1]

book = load_workbook(excel_path)
ws = book['Sheet']

# Add headers if the first cell in the extra columns is empty
ws.cell(row = 1, column = col_count + 1).value = 'N_H2S'
ws.cell(row = 1, column = col_count + 2).value = 'N_CO2'
ws.cell(row = 1, column = col_count + 3).value = 'S_CO2+H2S'
ws.cell(row = 1, column = col_count + 4).value = 'TSN_simu'

# Loop through txt lines
while i <= num_lines - 6:
    if 'best_mol_' in lines[i]:
        Structure_name = lines[i].strip()
        N_CH4 = lines[i + 1]
        N_C2 = lines[i + 2]
        N_C3 = lines[i + 3]
        N_CO2 = lines[i + 4]
        N_H2S = lines[i + 5]

        N1 = float(N_CH4.split()[5])  # CH4
        N2 = float(N_C2.split()[5])  # C2
        N3 = float(N_C3.split()[5])  # C3
        N4 = float(N_CO2.split()[5])  # CO2
        N5 = float(N_H2S.split()[5])  # H2S

        Structure_name_full = Structure_name + '.cif'
        N_total = N4 + N5

        # Avoid division by zero
        denom = N1 + N2 + N3
        numer = N4 + N5
        if denom == 0 or numer == 0:
            S_total = 0.0
        else:
            S_total = (numer / 0.15) / (denom / 0.85)

        TSN = N_total * np.log10(S_total) if S_total > 0 else 0.0

        # Write TSN, N4, N5, and S_total to Excel
        for j in range(row_count):
            name = df.iloc[j, 0]
            if name == Structure_name_full:
                ws.cell(row = j + 1, column = col_count + 1).value = N5 # H2S
                ws.cell(row = j + 1, column = col_count + 2).value = N4 # CO2
                ws.cell(row = j + 1, column = col_count + 3).value = S_total
                ws.cell(row = j + 1, column = col_count + 4).value = TSN
                break
        i += 6
        
    else:
        i += 1

# Save workbook once after the loop
book.save(excel_path)
