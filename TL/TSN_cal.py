# -*- coding: utf-8 -*-

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# Read txt file
with open('N.txt') as f:
    lines = f.readlines()
num_lines = len(lines)
i = 0

# Read Excel file
excel_path = 'TL_data_target_task_train.xlsx'
df = pd.read_excel(excel_path, header = None, engine = 'openpyxl')
row_count = len(df)

book = load_workbook(excel_path)
ws = book['Sheet']

# Find or create the four target columns
headers = [ws.cell(row = 1, column = c).value for c in range(1, ws.max_column + 1)]
target_headers = ['N_H2S', 'N_CO2', 'S_CO2+H2S', 'TSN_simu']

if all(h in headers for h in target_headers):
    # reuse existing columns
    col_NH2S = headers.index('N_H2S') + 1
    col_NCO2 = headers.index('N_CO2') + 1
    col_S    = headers.index('S_CO2+H2S') + 1
    col_TSN  = headers.index('TSN_simu') + 1
else:
    # append new columns to the right
    start_col = ws.max_column + 1
    col_NH2S = start_col
    col_NCO2 = start_col + 1
    col_S    = start_col + 2
    col_TSN  = start_col + 3

    ws.cell(row = 1, column = col_NH2S).value = 'N_H2S'
    ws.cell(row = 1, column = col_NCO2).value = 'N_CO2'
    ws.cell(row = 1, column = col_S).value    = 'S_CO2+H2S'
    ws.cell(row = 1, column = col_TSN).value  = 'TSN_simu'

# Parse lines
while i <= num_lines - 6:
    if 'best_mol_' in lines[i]:
        Structurename = lines[i].strip()
        N_CH4 = lines[i + 1]
        N_C2  = lines[i + 2]
        N_C3  = lines[i + 3]
        N_CO2 = lines[i + 4]
        N_H2S = lines[i + 5]

        N1 = float(N_CH4.split()[5])  # CH4
        N2 = float(N_C2.split()[5])   # C2
        N3 = float(N_C3.split()[5])   # C3
        N4 = float(N_CO2.split()[5])  # CO2
        N5 = float(N_H2S.split()[5])  # H2S

        Structurename_full = Structurename + '.cif'
        N_total = N4 + N5

        numer = N4 + N5
        denom = N1 + N2 + N3
        if numer <= 0 or denom <= 0:
            S_total = 0.0
            TSN = 0.0
        else:
            S_total = (numer / 0.15) / (denom / 0.85)
            TSN = N_total * np.log10(S_total) if S_total > 0 else 0.0

        # write into fixed columns
        for j in range(row_count):
            name = df.iloc[j, 0]
            if name == Structurename_full:
                ws.cell(row = j + 1, column = col_NH2S).value = N5   # H2S
                ws.cell(row = j + 1, column = col_NCO2).value = N4   # CO2
                ws.cell(row = j + 1, column = col_S).value    = S_total
                ws.cell(row = j + 1, column = col_TSN).value  = TSN
                break

        i += 6
    else:
        i += 1

book.save(excel_path)
