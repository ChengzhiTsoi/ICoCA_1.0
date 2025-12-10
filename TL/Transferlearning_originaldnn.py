# -*- coding: utf-8 -*-

from sklearn.model_selection import train_test_split
import pandas as pd
from sklearn import preprocessing
import torch.optim as optim
from ignite.engine import Engine, Events, create_supervised_evaluator
from ignite.metrics import Loss
from ignite.contrib.metrics.regression import R2Score
import torch, joblib
import torch.nn as nn
from torch.utils.data import Dataset
import numpy as np
from sklearn.preprocessing import StandardScaler
from openpyxl import load_workbook

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

# Load dataset (with header)
dataset_test_tt = pd.read_excel('TL_data_target_task_test.xlsx', header = 0, engine = 'openpyxl')
num_columns_test = len(dataset_test_tt.columns)
Structure_name = dataset_test_tt.iloc[:, 0].values

# Features: skip name col, exclude the last 4 derived/label cols
X_tt_test = dataset_test_tt.iloc[:, 1:num_columns_test-5].values

# NOTE: Ideally reuse the scaler fitted on TRAIN data.
scaler = joblib.load('TL/scaler_pretrained.pkl')
X_tt_test = scaler.transform(X_tt_test)

data_X_tt = torch.tensor(X_tt_test, dtype = torch.float32).to(device)

# Model
class NeuralNet(nn.Module):
    def __init__(self, input_size, hidden_size1, hidden_size2, hidden_size3, output_size):
        super(NeuralNet, self).__init__()
        self.input_layer = nn.Linear(input_size, hidden_size1)
        self.hidden_layer1 = nn.Linear(hidden_size1, hidden_size2)
        self.hidden_layer2 = nn.Linear(hidden_size2, hidden_size3)
        self.output_layer = nn.Linear(hidden_size3, output_size)
        self.acti = nn.ReLU()

    def forward(self, x):
        x = self.acti(self.input_layer(x))
        x = self.acti(self.hidden_layer1(x))
        x = self.acti(self.hidden_layer2(x))
        x = self.output_layer(x)
        return x

input_size = X_tt_test.shape[1]
model = NeuralNet(input_size, 64, 32, 16, 1)


# ===================== Prediction =====================
trained_model = torch.load('Pretrained_model.ckpt')
trained_model.to(device)
trained_model.eval()

with torch.no_grad():
    y_pred_tt = trained_model(data_X_tt).cpu().numpy().squeeze() # 1D array

# Write predictions to Excel (new column 'TSN_pred')
path_excel = 'TL_data_target_task_test.xlsx'
df = pd.read_excel(path_excel, header = 0, engine = 'openpyxl')
row_count = len(df)
column_count = len(df.columns)

book = load_workbook(path_excel)
ws = book['Sheet']

headers = list(df.columns)
if 'TSN_pred' in headers:
    pred_col = headers.index('TSN_pred') + 1
else:
    pred_col = column_count + 1
    ws.cell(row=1, column=pred_col).value = 'TSN_pred'

# Fast name -> row index mapping
name_to_row = {str(df.iloc[j, 0]).strip(): j for j in range(row_count)}

for name, pred in zip(Structure_name, y_pred_tt):
    key = str(name).strip()
    idx = name_to_row.get(key, None)
    if idx is not None:
        ws.cell(row = idx + 2, column = pred_col).value = float(pred) # +2: row 1 is header

book.save(path_excel)
print('The script Transferlearning_originaldnn.py has completed.')
