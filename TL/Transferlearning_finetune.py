# -*- coding: utf-8 -*-

from sklearn.model_selection import train_test_split
import pandas as pd
import torch.optim as optim
from ignite.engine import Engine, Events, create_supervised_evaluator
from ignite.metrics import Loss
from ignite.contrib.metrics.regression import R2Score
import torch
import torch.nn as nn
from torch.utils.data import Dataset
import numpy as np
from sklearn.preprocessing import StandardScaler
from openpyxl import load_workbook

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

# ===================== Transfer learning =====================
# Loading data
dataset_train_tt = pd.read_excel('TL_data_target_task_train.xlsx', header = 0, engine = 'openpyxl')
dataset_train_tt = dataset_train_tt.dropna()

num_columns_train = len(dataset_train_tt.columns)
# Features: skip name col (0), exclude last 4 derived/label cols
X_train_tt = dataset_train_tt.iloc[:, 1:num_columns_train-4].values
# Target by name (safer)
y_train_tt = dataset_train_tt['TSN_simu'].values

# Scale features (fit on train only)
scaler = StandardScaler()
scaler.fit(X_train_tt)
X_train_tt = scaler.transform(X_train_tt)
y_train_tt = y_train_tt.reshape(-1, 1)

X_train_tt_1, X_val_tt, y_train_tt_1, y_val_tt = train_test_split(
    X_train_tt, y_train_tt, test_size=0.1, random_state=42
)

train_set_tt = pd.concat([pd.DataFrame(X_train_tt_1), pd.DataFrame(y_train_tt_1)], axis = 1)
val_set_tt = pd.concat([pd.DataFrame(X_val_tt), pd.DataFrame(y_val_tt)], axis = 1)
n_features = X_train_tt_1.shape[1]

# Model (not used if you only fine-tune the loaded one, but kept for clarity)
class NeuralNet(nn.Module):
    def __init__(self, input_size, hidden_size1, hidden_size2, hidden_size3, output_size):
        super(NeuralNet, self).__init__()
        self.input_layer = nn.Linear(input_size, hidden_size1)
        self.hidden_layer1 = nn.Linear(hidden_size1, hidden_size2)
        self.hidden_layer2 = nn.Linear(hidden_size2, hidden_size3)
        self.output_layer = nn.Linear(hidden_size3, output_size)
        self.acti = nn.ReLU()

    def forward(self, x):
        x = self.acti(self.input_layer(x))
        x = self.acti(self.hidden_layer1(x))
        x = self.acti(self.hidden_layer2(x))
        x = self.output_layer(x)
        return x

# (Optional) Create a same-arch fresh model if needed
model = NeuralNet(n_features, 64, 32, 16, 1)

# Load the pretrained model you want to fine-tune
trained_model = torch.load('Pretrained_model.ckpt')
trained_model.to(device)
print(trained_model)

class MyDataset(Dataset):
    def __init__(self, dataframe, data_col_start, data_col_end, labels_col):
        x = dataframe.iloc[:, data_col_start:data_col_end].values
        y = dataframe.iloc[:, labels_col].values
        self.x = torch.tensor(x, dtype = torch.float32)
        self.y = torch.tensor(y, dtype = torch.float32).view(-1, 1)

    def __len__(self):
        return len(self.y)

    def __getitem__(self, index):
        return self.x[index], self.y[index]

train_set_1_tt = MyDataset(train_set_tt, 0, n_features, n_features)
val_set_1_tt = MyDataset(val_set_tt, 0, n_features, n_features)
batch_size = 32

# Dataloaders
train_loader_tt = torch.utils.data.DataLoader(train_set_1_tt, batch_size = batch_size, shuffle = True)
val_loader_tt = torch.utils.data.DataLoader(val_set_1_tt, batch_size = len(val_set_1_tt))

# Freeze/unfreeze
for p in trained_model.input_layer.parameters(): p.requires_grad = False
for p in trained_model.hidden_layer1.parameters(): p.requires_grad = True
for p in trained_model.hidden_layer2.parameters(): p.requires_grad = False
for p in trained_model.output_layer.parameters(): p.requires_grad = False

criterion = nn.MSELoss()
metrics_tt = {"loss": Loss(criterion), "r_2": R2Score()}
optimizer_tt = optim.Adam(filter(lambda p: p.requires_grad, trained_model.parameters()), lr = 1e-4)

# Training step (send batch to device)
def train_step_1(engine, batch):
    x, y = batch
    x = x.to(device)
    y = y.to(device)
    trained_model.train()
    optimizer_tt.zero_grad()
    y_pred = trained_model(x)
    loss = criterion(y_pred, y)
    loss.backward()
    optimizer_tt.step()
    return loss.item()

transfer_trainer = Engine(train_step_1)
train_evaluator = create_supervised_evaluator(trained_model, metrics = metrics_tt, device = device)
val_evaluator = create_supervised_evaluator(trained_model, metrics = metrics_tt, device = device)

train_loss_tt, train_r_2_tt, val_loss_tt, val_r_2_tt = [], [], [], []

@transfer_trainer.on(Events.EPOCH_COMPLETED(every=10))
def store_metrics(engine):
    train_evaluator.run(train_loader_tt)
    val_evaluator.run(val_loader_tt)
    out = train_evaluator.state.metrics
    out_2 = val_evaluator.state.metrics
    train_loss_tt.append(out["loss"])
    train_r_2_tt.append(out["r_2"])
    val_loss_tt.append(out_2["loss"])
    val_r_2_tt.append(out_2["r_2"])

transfer_trainer.run(train_loader_tt, max_epochs = 1000)

for name, metric in metrics_tt.items():
    metric.attach(transfer_trainer, name)


# ===================== Prediction =====================
trained_model.eval()

# Load test set
dataset_test_tt = pd.read_excel('TL_data_target_task_test.xlsx', header = 0, engine= 'openpyxl')
num_columns_test = len(dataset_test_tt.columns)
Structure_name = dataset_test_tt.iloc[:, 0].values

# Exclude last 4 derived/label columns just like train
X_tt_test = dataset_test_tt.iloc[:, 1:num_columns_test-4].values
X_tt_test = scaler.transform(X_tt_test)

data_X_tt = torch.tensor(X_tt_test, dtype=torch.float32).to(device)
with torch.no_grad():
    y_pred_tt = trained_model(data_X_tt).cpu().numpy().squeeze() # 1D array

# Write predictions to a NEW column 'TSN_pred'
path_excel = 'TL_data_target_task_test.xlsx'
df_test = pd.read_excel(path_excel, header = 0, engine = 'openpyxl')
row_count = len(df_test)
column_count = len(df_test.columns)

book = load_workbook(path_excel)
ws = book['Sheet']

pred_col = column_count + 1
ws.cell(row=1, column = pred_col).value = 'TSN_pred'

# Fast name -> row mapping
name_to_row = {str(df_test.iloc[j, 0]).strip(): j for j in range(row_count)}

for name, pred in zip(Structure_name, y_pred_tt):
    key = str(name).strip()
    if key in name_to_row:
        excel_row = name_to_row[key] + 2 # +2 because row 1 is header
        ws.cell(row = excel_row, column = pred_col).value = float(pred)

book.save(path_excel)

# ===================== Top 20% selection =====================
order = np.argsort(y_pred_tt)[::-1] # indices sorted by prediction desc
choose_number = max(1, int(round(0.2 * len(order))))
top_indices = order[:choose_number]
selected_names = [str(Structure_name[idx]) for idx in top_indices]
# drop ".cif" suffix if present
selected_names = [n[:-4] if n.lower().endswith('.cif') else n for n in selected_names]
pd.DataFrame(selected_names, columns = ['Name']).to_csv('Structure_verify_model.csv', index = False, encoding = 'utf-8')
print('The script Transferlearning_finetune.py has completed.')